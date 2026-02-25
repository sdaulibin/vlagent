"""
询证函批量识别测试脚本

批量处理指定目录下的 PDF 询证函文件，识别 12 项关键字段，并导出为 Excel。

用法:
    cd backend
    uv run python tests/test_batch_confirmation.py
"""
import os
import sys
import json
import time
import traceback
from datetime import datetime

# 添加项目根目录到 path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from services.pdf.pdf_utils import split_pdf_to_images
from services.core.request_ai import request_stream
from src.config import MODEL_LOCAL
from src.json_repair import fix_json
from src.confirmation_letter.service import (
    FIELD_EXTRACTION_PROMPT,
    _validate_and_normalize_fields,
    extract_fields_from_images,
    ALL_FIELDS,
    _check_format,
)

# ========== 配置 ==========
INPUT_DIR = "/Users/binginx/Desktop/2026年/星辰/运营管理部/50样本"
OUTPUT_DIR = INPUT_DIR  # Excel 输出到同目录
MAX_WORKERS = 1  # 串行处理，避免 AI 服务过载
DPI = 200
# ==========================

# 字段中文名映射
FIELD_LABELS = {
    "filename": "文件名",
    "confirmation_no": "函证编号",
    "accounting_firm": "事务所名称",
    "reply_address": "回函地址",
    "contact_person": "联系人",
    "phone": "电话",
    "postal_code": "邮编",
    "debit_account": "扣费账号",
    "cutoff_date": "截止日期",
    "start_date": "起始日期",
    "end_date": "终止日期",
    "seal_date": "印章日期",
    "seal_name": "印章名称",
    "signature_name": "落款名称",
    "status": "识别状态",
    "duration_s": "耗时(秒)",
    "error": "错误信息",
}

# 用于格式化打印的字段列表
PRINT_FIELD_LABELS = {
    "confirmation_no": "函证编号",
    "accounting_firm": "事务所名称",
    "reply_address": "回函地址",
    "contact_person": "联系人",
    "phone": "电话",
    "postal_code": "邮编",
    "debit_account": "扣费账号",
    "cutoff_date": "截止日期",
    "start_date": "起始日期",
    "end_date": "终止日期",
    "seal_date": "印章日期",
    "seal_name": "印章名称",
    "signature_name": "落款名称",
}

FIELD_KEYS = list(FIELD_LABELS.keys())


def _extract_fields_from_image(image_path: str) -> dict:
    """从单张图片提取字段"""
    response = request_stream(
        question=FIELD_EXTRACTION_PROMPT,
        file_base=image_path,
        model=MODEL_LOCAL,
        show_request=False,
    ).strip()
    try:
        return json.loads(fix_json(response))
    except Exception as e:
        print(f"JSON 解析失败: {e}, 原始响应: {response[:200]}")
        return {}


def recognize_single_pdf(pdf_path: str) -> dict:
    """识别单个 PDF 询证函（使用优化后的单次多图 AI 调用）"""
    filename = os.path.basename(pdf_path)
    result = {"filename": filename, "status": "failed", "error": "", "duration_s": 0}

    start = time.time()
    try:
        # 1. PDF 转图片
        tmp_dir = os.path.join(os.path.dirname(pdf_path), "_tmp_images")
        os.makedirs(tmp_dir, exist_ok=True)

        image_paths = split_pdf_to_images(pdf_path, tmp_dir, dpi=DPI)
        if not image_paths:
            result["error"] = "PDF 转图片失败"
            return result

        # 2. 所有页面一次性提交 AI
        fields = extract_fields_from_images(image_paths)
        merged_text = fields.pop("raw_text", "")

        # 3. 后处理
        normalized = _validate_and_normalize_fields(fields, merged_text)
        result.update(normalized)
        result["status"] = "success"

        # 4. 清理临时图片
        for p in image_paths:
            if os.path.exists(p):
                os.remove(p)

    except Exception as e:
        result["error"] = str(e)
        traceback.print_exc()

    result["duration_s"] = round(time.time() - start, 1)
    return result


def print_recognition_result(result: dict):
    """格式化打印识别结果"""
    filename = result.get("filename", "")
    duration = result.get("duration_s", 0)
    print(f"\n  {'=' * 50}")
    print(f"  文件: {filename}  耗时: {duration}s")
    print(f"  {'-' * 50}")
    for key, label in PRINT_FIELD_LABELS.items():
        value = result.get(key, "")
        print(f"    {label}: {value}")
    print(f"  {'-' * 50}")
    print(f"  原始 JSON:")
    # 只输出 13 个字段
    fields_json = {k: result.get(k, "") for k in PRINT_FIELD_LABELS}
    print(f"  {json.dumps(fields_json, ensure_ascii=False, indent=4)}")
    print(f"  {'=' * 50}")


def export_to_excel(results: list, output_path: str):
    """导出识别结果到 Excel"""
    try:
        import openpyxl
    except ImportError:
        print("正在安装 openpyxl...")
        os.system(f"{sys.executable} -m pip install openpyxl")
        import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "询证函识别结果"

    # 表头
    headers = [FIELD_LABELS[k] for k in FIELD_KEYS]
    ws.append(headers)

    # 设置表头样式
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    for r in results:
        row = [r.get(k, "") for k in FIELD_KEYS]
        ws.append(row)

    # 自动列宽
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val.encode("utf-8")))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    wb.save(output_path)
    print(f"\n✅ Excel 已导出: {output_path}")


def main():
    print("=" * 60)
    print("  银行询证函批量识别测试")
    print(f"  输入目录: {INPUT_DIR}")
    print(f"  开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # 扫描 PDF 文件
    pdf_files = sorted([
        os.path.join(INPUT_DIR, f)
        for f in os.listdir(INPUT_DIR)
        if f.lower().endswith(".pdf")
    ])

    total = len(pdf_files)
    print(f"\n找到 {total} 个 PDF 文件\n")

    if total == 0:
        print("没有找到 PDF 文件，退出。")
        return

    # 逐个处理
    results = []
    success_count = 0
    total_start = time.time()

    for idx, pdf_path in enumerate(pdf_files, 1):
        filename = os.path.basename(pdf_path)
        print(f"[{idx}/{total}] 正在处理: {filename} ...", end=" ", flush=True)

        result = recognize_single_pdf(pdf_path)
        results.append(result)

        if result["status"] == "success":
            success_count += 1
            print(f"✅ ({result['duration_s']}s)")
            print_recognition_result(result)
        else:
            print(f"❌ {result['error']}")

    total_duration = round(time.time() - total_start, 1)

    # 导出 Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(OUTPUT_DIR, f"询证函识别结果_{timestamp}.xlsx")
    export_to_excel(results, excel_path)

    # 统计
    print(f"\n{'=' * 60}")
    print(f"  处理完成!")
    print(f"  总数: {total}  成功: {success_count}  失败: {total - success_count}")
    print(f"  总耗时: {total_duration}s  平均: {round(total_duration / total, 1)}s/份")
    print(f"  结果文件: {excel_path}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
