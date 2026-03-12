"""
法律服务联络人信息提取批量测试脚本

批量处理指定目录下的 PDF 文件，识别联络人信息表中的四项关键字段，并导出为 Excel。

用法:
    cd backend
    uv run python tests/test_batch_legal_contact.py
"""
import os
import sys
import json
import time
import traceback
from datetime import datetime

# 添加项目根目录到 path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.legal_contact.service import process_legal_contact

# ========== 配置 ==========
INPUT_DIR = "/Users/binginx/Downloads/联络人测试"  # 请修改为真实的目录
OUTPUT_DIR = INPUT_DIR  # Excel 输出到同目录
# ==========================

# 字段中文名映射 (按照图2和要求定制)
FIELD_LABELS = {
    "filename": "文件名",
    "law_firm_name": "律所名称",
    "name_and_phone": "联系人",
    "email": "邮箱",
    "status": "识别状态",
    "duration_s": "耗时(秒)",
    "error": "错误信息",
}

# 导出 Excel 时的列序
EXPORT_FIELDS = [
    "filename",
    "law_firm_name",
    "name_and_phone",
    "email",
    "status",
    "duration_s",
    "error"
]

# 用于格式化打印的字段列表
PRINT_FIELD_LABELS = {
    "law_firm_name": "律所名称",
    "name_and_phone": "联系人",
    "email": "邮箱",
}


def recognize_single_pdf(pdf_path: str) -> dict:
    """识别单个 PDF（联络人信息表）"""
    filename = os.path.basename(pdf_path)
    result = {"filename": filename, "status": "failed", "error": "", "duration_s": 0}

    start = time.time()
    try:
        # 调用核心服务进行提取
        extracted = process_legal_contact(pdf_path)
        result.update(extracted)
        result["status"] = "success"
    except Exception as e:
        result["error"] = str(e)
        traceback.print_exc()

    result["duration_s"] = round(time.time() - start, 1)
    return result


def print_recognition_result(result: dict):
    """格式化打印识别结果"""
    filename = result.get("filename", "")
    duration = result.get("duration_s", 0)
    print(f"\n  {'=' * 50}")
    print(f"  文件: {filename}  耗时: {duration}s")
    print(f"  {'-' * 50}")
    for key, label in PRINT_FIELD_LABELS.items():
        value = result.get(key, "")
        print(f"    {label}: {value}")
    print(f"  {'-' * 50}")
    # 输出 JSON 部分
    fields_json = {k: result.get(k, "") for k in PRINT_FIELD_LABELS}
    print(f"  原始 JSON:\n  {json.dumps(fields_json, ensure_ascii=False, indent=4)}")
    print(f"  {'=' * 50}")


def export_to_excel(results: list, output_path: str):
    """导出识别结果到 Excel"""
    try:
        import openpyxl
    except ImportError:
        print("正在安装 openpyxl...")
        os.system(f"{sys.executable} -m pip install openpyxl")
        import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "联络人信息"

    # 表头
    headers = [FIELD_LABELS[k] for k in EXPORT_FIELDS]
    ws.append(headers)

    # 设置表头样式
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True)
    # 使用淡淡的绿色来模拟图2样式
    header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    for r in results:
        row = []
        for k in EXPORT_FIELDS:
            val = r.get(k, "")
            # 对于核心字段，如果为空则填充“未识别”，符合图2形式
            if not val and k in ["law_firm_name", "name_and_phone", "email"]:
                val = "未识别"
            row.append(val)
        ws.append(row)

    # 允许单元格换行展示并设置数据行样式
    data_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = data_fill
            cell.alignment = Alignment(horizontal="left", wrap_text=True)

    # 自动列宽
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val.encode("utf-8")))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    wb.save(output_path)


def main():
    print("=" * 60)
    print("  法律服务联络人信息批量识别测试")
    print(f"  输入目录: {INPUT_DIR}")
    print(f"  开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    if not os.path.exists(INPUT_DIR):
        print(f"【注意】输入目录不存在: {INPUT_DIR}")
        print("请打开 tests/test_batch_legal_contact.py 文件，修改 INPUT_DIR 为真实包含 PDF 的目录。")
        return

    # 扫描 PDF 文件
    pdf_files = sorted([
        os.path.join(INPUT_DIR, f)
        for f in os.listdir(INPUT_DIR)
        if f.lower().endswith(".pdf")
    ])

    total = len(pdf_files)
    print(f"\n找到 {total} 个 PDF 文件\n")

    if total == 0:
        print("没有找到 PDF 文件，退出。")
        return

    # 初始化和实时写 Excel 的准备
    results = []
    success_count = 0
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(OUTPUT_DIR, f"联络人识别结果_{timestamp}.xlsx")

    total_start = time.time()
    
    # 逐个串行处理
    for idx, pdf_path in enumerate(pdf_files, 1):
        filename = os.path.basename(pdf_path)
        print(f"[{idx}/{total}] 正在处理: {filename} ...", end=" ", flush=True)

        result = recognize_single_pdf(pdf_path)
        results.append(result)

        if result["status"] == "success":
            success_count += 1
            print(f"✅ ({result['duration_s']}s)")
            print_recognition_result(result)
        else:
            print(f"❌ {result['error']}")
            
        # 实时保存到 Excel
        export_to_excel(results, excel_path)

    total_duration = round(time.time() - total_start, 1)

    # 统计
    print(f"\n{'=' * 60}")
    print(f"  处理完成!")
    print(f"  总数: {total}  成功: {success_count}  失败: {total - success_count}")
    print(f"  总耗时: {total_duration}s  平均: {round(total_duration / total, 1)}s/份")
    print(f"  结果文件: {excel_path}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
