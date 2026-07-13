from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

from src.credit_comparison.core.dto import ExcelProfitLossRecord
from src.credit_comparison.core.text_utils import normalize_indicator_name

HEADER_ROW_COUNT = 4
HEADER_SCAN_LIMIT = 12
PURE_DIGIT_SHEET_PATTERN = re.compile(r"^\d+$")
LEADING_LETTER_PATTERN = re.compile(r"^\s*([A-Za-z]+)")
logger = logging.getLogger(__name__)

FIELD_BY_SEMANTIC = {
    ("cur", "rmb", "balance"): "cur_rmb_balance",
    ("cur", "rmb", "occur"): "cur_rmb_occur",
    ("cur", "foreign", "balance"): "cur_foreign_balance",
    ("cur", "foreign", "occur"): "cur_foreign_occur",
    ("cur", "usd_total", "balance"): "cur_foreign_total_balance",
    ("cur", "usd_total", "occur"): "cur_foreign_total_occur",
    ("pre", "rmb", "balance"): "pre_rmb_balance",
    ("pre", "rmb", "occur"): "pre_rmb_occur",
    ("pre", "foreign", "balance"): "pre_foreign_balance",
    ("pre", "foreign", "occur"): "pre_foreign_occur",
    ("pre", "usd_total", "balance"): "pre_foreign_total_balance",
    ("pre", "usd_total", "occur"): "pre_foreign_total_occur",
}


class ExcelSheetAdapter:
    """统一封装不同 Excel 库的工作表读取接口。"""

    def __init__(self, name: str, nrows: int, ncols: int, merged_cells: list[tuple[int, int, int, int]]) -> None:
        self.name = name
        self.nrows = nrows
        self.ncols = ncols
        self.merged_cells = merged_cells

    def cell_value(self, row_index: int, col_index: int) -> Any:
        raise NotImplementedError


class XlrdSheetAdapter(ExcelSheetAdapter):
    """xlrd 工作表适配器。"""

    def __init__(self, sheet_obj: Any) -> None:
        self._sheet_obj = sheet_obj
        super().__init__(
            name=str(sheet_obj.name or ""),
            nrows=int(sheet_obj.nrows or 0),
            ncols=int(sheet_obj.ncols or 0),
            merged_cells=list(getattr(sheet_obj, "merged_cells", [])),
        )

    def cell_value(self, row_index: int, col_index: int) -> Any:
        return self._sheet_obj.cell_value(row_index, col_index)


class OpenpyxlSheetAdapter(ExcelSheetAdapter):
    """openpyxl 工作表适配器。"""

    def __init__(self, sheet_obj: Any) -> None:
        self._sheet_obj = sheet_obj
        merged_cells = [
            (merged_range.min_row - 1, merged_range.max_row, merged_range.min_col - 1, merged_range.max_col)
            for merged_range in sheet_obj.merged_cells.ranges
        ]
        super().__init__(
            name=str(sheet_obj.title or ""),
            nrows=int(sheet_obj.max_row or 0),
            ncols=int(sheet_obj.max_column or 0),
            merged_cells=merged_cells,
        )

    def cell_value(self, row_index: int, col_index: int) -> Any:
        return self._sheet_obj.cell(row=row_index + 1, column=col_index + 1).value


class ExcelWorkbookAdapter:
    """统一封装不同 Excel 库的工作簿接口。"""

    def sheet_names(self) -> list[str]:
        raise NotImplementedError

    def sheet_by_name(self, sheet_name: str) -> ExcelSheetAdapter:
        raise NotImplementedError


class XlrdWorkbookAdapter(ExcelWorkbookAdapter):
    """xlrd 工作簿适配器。"""

    def __init__(self, workbook: Any) -> None:
        self._workbook = workbook

    def sheet_names(self) -> list[str]:
        return [str(name or "") for name in self._workbook.sheet_names()]

    def sheet_by_name(self, sheet_name: str) -> ExcelSheetAdapter:
        return XlrdSheetAdapter(self._workbook.sheet_by_name(sheet_name))


class OpenpyxlWorkbookAdapter(ExcelWorkbookAdapter):
    """openpyxl 工作簿适配器。"""

    def __init__(self, workbook: Any) -> None:
        self._workbook = workbook

    def sheet_names(self) -> list[str]:
        return [str(name or "") for name in self._workbook.sheetnames]

    def sheet_by_name(self, sheet_name: str) -> ExcelSheetAdapter:
        return OpenpyxlSheetAdapter(self._workbook[sheet_name])


def normalize_excel_cell(value: Any) -> Any:
    """标准化 Excel 单元格值。"""

    if isinstance(value, str):
        return value.strip()
    return value


def normalize_header_token(text: str) -> str:
    """标准化表头文本，便于按关键词识别语义。"""

    return text.replace("\n", "").replace("\r", "").replace(" ", "").strip()


def is_non_empty_cell(value: Any) -> bool:
    """判断单元格是否为非空值。"""

    normalized = normalize_excel_cell(value)
    return normalized not in ("", None)


def format_excel_code(value: Any) -> str:
    """标准化指标代码文本。"""

    value = normalize_excel_cell(value)
    if value in ("", None):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_excel_number(value: Any) -> float | None:
    """标准化 Excel 数值。"""

    value = normalize_excel_cell(value)
    if value in ("", None):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def extract_sheet_letter_prefix(sheet_name: str) -> str:
    """从表单名中提取前导字母前缀。"""

    match = LEADING_LETTER_PATTERN.match(str(sheet_name or "").strip())
    return match.group(1) if match else ""


def normalize_excel_sheet_name(sheet_name: str, default_prefix: str) -> str:
    """标准化 Excel 表单名称。

    规则：
    - 若当前表单名只有数字，则使用第一张表单名的字母前缀进行拼接
    - 否则保持原样
    """

    normalized_name = str(sheet_name or "").strip()
    if not normalized_name:
        return ""
    if PURE_DIGIT_SHEET_PATTERN.fullmatch(normalized_name) and default_prefix:
        return f"{default_prefix}{normalized_name}"
    return normalized_name


def summarize_row_values(row_values: list[Any], max_items: int = 8) -> list[str]:
    """提取一行中的非空单元格预览，便于日志观察。"""

    preview: list[str] = []
    for col_index, value in enumerate(row_values):
        normalized = normalize_excel_cell(value)
        if normalized in ("", None):
            continue
        preview.append(f"{col_index}:{normalized}")
        if len(preview) >= max_items:
            break
    return preview


def detect_header_rows(sheet_obj, sheet_name: str) -> list[int]:
    """在工作表前部扫描前 4 个非空表头行。

    这里不再强依赖“物理前 4 行”必须连续，
    而是允许表头区域中穿插空行。
    """

    header_rows: list[int] = []
    scan_limit = min(sheet_obj.nrows, HEADER_SCAN_LIMIT)
    logger.debug(
        "开始扫描表头行: sheet=%s, nrows=%s, ncols=%s, scan_limit=%s",
        sheet_name,
        sheet_obj.nrows,
        sheet_obj.ncols,
        scan_limit,
    )
    for row_index in range(scan_limit):
        row_values = [sheet_obj.cell_value(row_index, col_index) for col_index in range(sheet_obj.ncols)]
        row_preview = summarize_row_values(row_values)
        is_non_empty = any(is_non_empty_cell(value) for value in row_values)
        logger.debug(
            "扫描行: sheet=%s, row=%s, non_empty=%s, preview=%s",
            sheet_name,
            row_index + 1,
            is_non_empty,
            row_preview,
        )
        if is_non_empty:
            header_rows.append(row_index)
            logger.debug(
                "识别到表头候选行: sheet=%s, row=%s, header_rows=%s",
                sheet_name,
                row_index + 1,
                [index + 1 for index in header_rows],
            )
        if len(header_rows) == HEADER_ROW_COUNT:
            break
    logger.debug("表头扫描完成: sheet=%s, header_rows=%s", sheet_name, [index + 1 for index in header_rows])
    return header_rows


def build_merged_value_lookup(sheet_obj, header_rows: list[int], sheet_name: str) -> dict[tuple[int, int], Any]:
    """展开表头区域合并单元格，统一取合并区域左上角值。"""

    lookup: dict[tuple[int, int], Any] = {}
    header_row_set = set(header_rows)
    merged_cells = getattr(sheet_obj, "merged_cells", [])
    logger.debug("读取工作表合并单元格信息: sheet=%s, raw_merged_cell_count=%s", sheet_name, len(merged_cells))
    for row_low, row_high, col_low, col_high in merged_cells:
        if not any(row_index in header_row_set for row_index in range(row_low, row_high)):
            continue
        top_left_value = normalize_excel_cell(sheet_obj.cell_value(row_low, col_low))
        logger.debug(
            "展开表头合并单元格: sheet=%s, rows=%s-%s, cols=%s-%s, top_left_value=%s",
            sheet_name,
            row_low + 1,
            row_high,
            col_low,
            col_high - 1,
            top_left_value,
        )
        for row_index in range(row_low, row_high):
            if row_index not in header_row_set:
                continue
            for col_index in range(col_low, col_high):
                lookup[(row_index, col_index)] = top_left_value
    logger.debug("表头合并单元格展开完成: sheet=%s, merged_cell_count=%s", sheet_name, len(lookup))
    return lookup


def get_header_cell_text(sheet_obj, merged_lookup: dict[tuple[int, int], Any], row_index: int, col_index: int) -> str:
    """读取表头单元格文本，优先使用合并单元格展开值。"""

    value = merged_lookup.get((row_index, col_index), sheet_obj.cell_value(row_index, col_index))
    normalized = normalize_excel_cell(value)
    return "" if normalized in ("", None) else str(normalized).strip()


def build_column_header_text(
    sheet_obj,
    merged_lookup: dict[tuple[int, int], Any],
    col_index: int,
    semantic_rows: list[int],
) -> str:
    """构造单列的多行表头文本。"""

    parts: list[str] = []
    for row_index in semantic_rows:
        text = get_header_cell_text(sheet_obj, merged_lookup, row_index, col_index)
        if text:
            parts.append(normalize_header_token(text))
    return " ".join(parts)


def build_column_header_tokens(
    sheet_obj,
    merged_lookup: dict[tuple[int, int], Any],
    col_index: int,
    semantic_rows: list[int],
) -> list[str]:
    """构造单列表头的分层 token 列表（按行顺序）。"""

    tokens: list[str] = []
    for row_index in semantic_rows:
        text = get_header_cell_text(sheet_obj, merged_lookup, row_index, col_index)
        normalized = normalize_header_token(text)
        tokens.append(normalized if normalized else "")
    return tokens


def detect_period(header_text: str) -> str | None:
    """识别期间维度。"""

    if "本期" in header_text:
        return "cur"
    if "上期" in header_text:
        return "pre"
    return None


def detect_period_from_tokens(header_tokens: list[str]) -> str | None:
    """按表头分层识别期间，优先使用最靠下且不歧义的 token。"""

    candidate_by_row: list[tuple[int, str]] = []
    for row_index, token in enumerate(header_tokens):
        if not token:
            continue
        period_hits: list[str] = []
        if "本期" in token:
            period_hits.append("cur")
        if "上期" in token:
            period_hits.append("pre")
        if len(period_hits) == 1:
            candidate_by_row.append((row_index, period_hits[0]))

    if not candidate_by_row:
        return None
    return sorted(candidate_by_row, key=lambda item: item[0], reverse=True)[0][1]


def detect_scope(header_text: str) -> str | None:
    """识别币种/口径维度。"""

    if "本外币" in header_text:
        return "foreign"
    if "美元合计" in header_text:
        return "usd_total"
    if "人民币" in header_text:
        return "rmb"
    return None


def detect_scope_from_tokens(header_tokens: list[str]) -> str | None:
    """按表头分层识别币种/口径，避免拼接后多口径串扰。"""

    candidate_by_row: list[tuple[int, str]] = []
    for row_index, token in enumerate(header_tokens):
        if not token:
            continue
        scope_hits: list[str] = []
        if "本外币" in token:
            scope_hits.append("foreign")
        if "美元合计" in token:
            scope_hits.append("usd_total")
        if "人民币" in token:
            scope_hits.append("rmb")
        if len(scope_hits) == 1:
            candidate_by_row.append((row_index, scope_hits[0]))

    if not candidate_by_row:
        return None
    # 更靠下的行通常更接近最终语义层，优先采用。
    return sorted(candidate_by_row, key=lambda item: item[0], reverse=True)[0][1]


def detect_measure(header_text: str) -> str | None:
    """识别余额/发生额维度。"""

    if "发生额" in header_text:
        return "occur"
    if "余额" in header_text:
        return "balance"
    return None


def detect_measure_from_tokens(header_tokens: list[str]) -> str | None:
    """按表头分层识别余额/发生额，避免拼接误判。"""

    candidate_by_row: list[tuple[int, str]] = []
    for row_index, token in enumerate(header_tokens):
        if not token:
            continue
        measure_hits: list[str] = []
        if "发生额" in token:
            measure_hits.append("occur")
        if "余额" in token:
            measure_hits.append("balance")
        if len(measure_hits) == 1:
            candidate_by_row.append((row_index, measure_hits[0]))

    if not candidate_by_row:
        return None
    return sorted(candidate_by_row, key=lambda item: item[0], reverse=True)[0][1]


def detect_sheet_columns(sheet_obj, sheet_name: str) -> dict[str, int]:
    """基于前 4 个非空表头行识别关键列位置。"""

    if sheet_obj.nrows <= 0:
        logger.warning("工作表为空，跳过列识别: sheet=%s", sheet_name)
        return {}

    header_rows = detect_header_rows(sheet_obj, sheet_name)
    if len(header_rows) < HEADER_ROW_COUNT:
        logger.warning(
            "表头行不足，无法识别表结构: sheet=%s, header_rows=%s",
            sheet_name,
            [index + 1 for index in header_rows],
        )
        return {}

    merged_lookup = build_merged_value_lookup(sheet_obj, header_rows, sheet_name)
    semantic_rows = header_rows[1:]
    column_map: dict[str, int] = {"header_end_row_index": header_rows[-1]}
    logger.debug(
        "开始识别表格结构: sheet=%s, header_rows=%s, semantic_rows=%s",
        sheet_name,
        [index + 1 for index in header_rows],
        [index + 1 for index in semantic_rows],
    )

    for col_index in range(sheet_obj.ncols):
        header_text = build_column_header_text(sheet_obj, merged_lookup, col_index, semantic_rows)
        compact_text = header_text.replace(" ", "")
        header_tokens = build_column_header_tokens(sheet_obj, merged_lookup, col_index, semantic_rows)
        if not compact_text:
            logger.debug("列为空白表头，跳过: sheet=%s, col=%s, tokens=%s", sheet_name, col_index, header_tokens)
            continue

        if "指标编号" in compact_text or "指标代码" in compact_text:
            column_map["code"] = col_index
            logger.debug(
                "识别到指标代码列: sheet=%s, col=%s, header_text=%s, tokens=%s",
                sheet_name,
                col_index,
                compact_text,
                header_tokens,
            )
            continue
        if "指标名称" in compact_text:
            column_map["name"] = col_index
            logger.debug(
                "识别到指标名称列: sheet=%s, col=%s, header_text=%s, tokens=%s",
                sheet_name,
                col_index,
                compact_text,
                header_tokens,
            )
            continue

        period = detect_period_from_tokens(header_tokens) or detect_period(compact_text)
        scope = detect_scope_from_tokens(header_tokens) or detect_scope(compact_text)
        measure = detect_measure_from_tokens(header_tokens) or detect_measure(compact_text)
        logger.debug(
            "分析列表头语义: sheet=%s, col=%s, header_text=%s, tokens=%s, period=%s, scope=%s, measure=%s",
            sheet_name,
            col_index,
            compact_text,
            header_tokens,
            period,
            scope,
            measure,
        )
        if not period or not scope or not measure:
            continue

        mapped_field = FIELD_BY_SEMANTIC.get((period, scope, measure))
        if mapped_field:
            column_map[mapped_field] = col_index
            logger.debug(
                "映射业务字段成功: sheet=%s, col=%s, mapped_field=%s",
                sheet_name,
                col_index,
                mapped_field,
            )

    if {"code", "name"}.issubset(column_map):
        logger.info("表格结构识别完成: sheet=%s, column_map=%s", sheet_name, column_map)
        return column_map
    logger.warning("表格结构识别失败，缺少关键列: sheet=%s, column_map=%s", sheet_name, column_map)
    return {}


def parse_excel_row(
    row_data: list[Any],
    column_map: dict[str, int],
    file_name: str,
    sheet_name: str,
    batch_id: str,
    row_index: int,
) -> dict | None:
    """解析单行 Excel 数据。"""

    code = format_excel_code(row_data[column_map["code"]])
    name = normalize_excel_cell(row_data[column_map["name"]])
    normalized_name = normalize_indicator_name(str(name)) if name not in ("", None) else ""
    if not code or not name or not normalized_name:
        logger.debug(
            "跳过数据行: file=%s, sheet=%s, row=%s, code=%s, name=%s, normalized_name=%s",
            file_name,
            sheet_name,
            row_index + 1,
            code,
            name,
            normalized_name,
        )
        return None

    record = ExcelProfitLossRecord(
        sheet=sheet_name.strip(),
        code=code,
        name=str(name).strip(),
        cur_rmb_balance=parse_excel_number(row_data[column_map["cur_rmb_balance"]]) if "cur_rmb_balance" in column_map else None,
        cur_rmb_occur=parse_excel_number(row_data[column_map["cur_rmb_occur"]]) if "cur_rmb_occur" in column_map else None,
        cur_foreign_balance=parse_excel_number(row_data[column_map["cur_foreign_balance"]]) if "cur_foreign_balance" in column_map else None,
        cur_foreign_occur=parse_excel_number(row_data[column_map["cur_foreign_occur"]]) if "cur_foreign_occur" in column_map else None,
        cur_foreign_total_balance=parse_excel_number(row_data[column_map["cur_foreign_total_balance"]]) if "cur_foreign_total_balance" in column_map else None,
        cur_foreign_total_occur=parse_excel_number(row_data[column_map["cur_foreign_total_occur"]]) if "cur_foreign_total_occur" in column_map else None,
        pre_rmb_balance=parse_excel_number(row_data[column_map["pre_rmb_balance"]]) if "pre_rmb_balance" in column_map else None,
        pre_rmb_occur=parse_excel_number(row_data[column_map["pre_rmb_occur"]]) if "pre_rmb_occur" in column_map else None,
        pre_foreign_balance=parse_excel_number(row_data[column_map["pre_foreign_balance"]]) if "pre_foreign_balance" in column_map else None,
        pre_foreign_occur=parse_excel_number(row_data[column_map["pre_foreign_occur"]]) if "pre_foreign_occur" in column_map else None,
        pre_foreign_total_balance=parse_excel_number(row_data[column_map["pre_foreign_total_balance"]]) if "pre_foreign_total_balance" in column_map else None,
        pre_foreign_total_occur=parse_excel_number(row_data[column_map["pre_foreign_total_occur"]]) if "pre_foreign_total_occur" in column_map else None,
        excel_row_index=row_index + 1,
        file_name=file_name,
        batch_id=batch_id,
    )
    db_record = record.to_db_dict()
    logger.debug(
        "生成入库记录: file=%s, sheet=%s, row=%s, code=%s, name=%s, record=%s",
        file_name,
        sheet_name,
        row_index + 1,
        code,
        str(name).strip(),
        db_record,
    )
    return db_record


def load_excel_workbook(file_path: str) -> ExcelWorkbookAdapter:
    """按后缀加载 Excel 工作簿。"""

    suffix = Path(file_path).suffix.lower()
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("缺少 openpyxl 依赖，无法解析 xlsx") from exc

        workbook = load_workbook(file_path, data_only=True)
        return OpenpyxlWorkbookAdapter(workbook)

    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("缺少 xlrd==1.2.0 依赖，无法解析 xls") from exc

    workbook = xlrd.open_workbook(file_path, formatting_info=True)
    return XlrdWorkbookAdapter(workbook)


def parse_excel_file(file_path: str, file_name: str, batch_id: str) -> list[dict]:
    """解析 Excel 文件。

    这里与 Word 解析保持一致，采用“单文件内存聚合”策略：
    - 当前函数只负责读取一个 Excel 文件并返回记录列表
    - 不在逐行解析时直接写数据库
    - 由上层 service 在整个文件解析完成后统一批量写库
    """

    logger.info("开始解析 Excel 文件: file_path=%s, file_name=%s, batch_id=%s", file_path, file_name, batch_id)
    workbook = load_excel_workbook(file_path)
    logger.info("打开 Excel 工作簿成功: file_name=%s, loader=%s", file_name, workbook.__class__.__name__)
    sheet_names = workbook.sheet_names()
    default_sheet_prefix = extract_sheet_letter_prefix(sheet_names[0]) if sheet_names else ""
    logger.info(
        "Excel 文件已打开: file_name=%s, sheet_names=%s, default_sheet_prefix=%s",
        file_name,
        sheet_names,
        default_sheet_prefix,
    )
    # 单文件内存聚合，等整个 Excel 文件解析完成后再由 service 批量写库。
    records: list[dict] = []
    for raw_sheet_name in sheet_names:
        normalized_sheet_name = normalize_excel_sheet_name(raw_sheet_name, default_sheet_prefix)
        logger.debug(
            "开始处理工作表: file_name=%s, raw_sheet_name=%s, normalized_sheet_name=%s",
            file_name,
            raw_sheet_name,
            normalized_sheet_name,
        )
        sheet_obj = workbook.sheet_by_name(raw_sheet_name)
        logger.debug(
            "工作表基础信息: file_name=%s, sheet=%s, nrows=%s, ncols=%s",
            file_name,
            normalized_sheet_name,
            sheet_obj.nrows,
            sheet_obj.ncols,
        )
        column_map = detect_sheet_columns(sheet_obj, normalized_sheet_name)
        if not column_map:
            logger.warning(
                "工作表未识别到可解析结构，跳过: file_name=%s, sheet=%s",
                file_name,
                normalized_sheet_name,
            )
            continue
        start_row = int(column_map.get("header_end_row_index", HEADER_ROW_COUNT - 1)) + 1
        logger.debug(
            "确定数据起始行: file_name=%s, sheet=%s, header_end_row=%s, start_row=%s, column_map=%s",
            file_name,
            normalized_sheet_name,
            int(column_map.get("header_end_row_index", HEADER_ROW_COUNT - 1)) + 1,
            start_row + 1,
            column_map,
        )
        sheet_record_count = 0
        for row_index in range(start_row, sheet_obj.nrows):
            row_data = [sheet_obj.cell_value(row_index, col) for col in range(sheet_obj.ncols)]
            logger.debug(
                "读取数据行: file_name=%s, sheet=%s, row=%s, preview=%s",
                file_name,
                normalized_sheet_name,
                row_index + 1,
                summarize_row_values(row_data),
            )
            record = parse_excel_row(row_data, column_map, file_name, normalized_sheet_name, batch_id, row_index)
            if record:
                records.append(record)
                sheet_record_count += 1
                logger.debug(
                    "记录加入批量入库列表: file_name=%s, sheet=%s, row=%s, sheet_record_count=%s, total_record_count=%s",
                    file_name,
                    normalized_sheet_name,
                    row_index + 1,
                    sheet_record_count,
                    len(records),
                )
        logger.info(
            "工作表解析完成: file_name=%s, sheet=%s, sheet_record_count=%s",
            file_name,
            normalized_sheet_name,
            sheet_record_count,
        )
    logger.info("Excel 文件解析完成，返回入库结果: file_name=%s, total_records=%s", file_name, len(records))
    return records


def get_excel_file_name(file_path: str) -> str:
    """返回 Excel 文件名。"""

    return Path(file_path).name
