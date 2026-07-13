"""
信用金额对账 - 异常记录导出（xlsx）。

与旧 ExceptionExportService 的数据加工逻辑完全一致，仅把同步 repository 调用
替换为异步 repository 调用。
"""
from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio.session import AsyncSession

from src.credit_comparison import repository
from src.credit_comparison.core.enums import ExceptionType
from src.credit_comparison.models import CreditCompareTask

EXCEPTION_DISPLAY_NAME_MAP = {
    "指标代码异常": "指标代码未找到",
    "指标名称异常": "指标名称不匹配",
    "指标数值异常": "指标金额计算有误",
    "表单无对应异常": "无关联表单",
    "关联公司数值异常": "关联公司增减金额不一致",
    "关联公司方向不一致": "关联公司增减方向与当前主句不一致",
    "关联公司格式异常": "格式异常",
    "同一记录关联公司重复出现": "同一记录关联公司重复出现",
    "余额缺失异常": "余额缺失",
    "计算要求异常": "无合适计算币种",
    "标点符号异常": "格式异常",
    "excel异常": "对应多条excel记录",
}
EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _display_exception_name(name: object) -> str:
    text = str(name or "").strip()
    return EXCEPTION_DISPLAY_NAME_MAP.get(text, text)


def _is_company_exception(exception_id: object, field_name: object = "") -> bool:
    try:
        exception_id_int = int(exception_id or 0)
    except (TypeError, ValueError):
        return False
    if exception_id_int in {
        int(ExceptionType.COMPANY_AMOUNT_ERROR),
        int(ExceptionType.COMPANY_DIRECTION_ERROR),
        int(ExceptionType.COMPANY_FORMAT_ERROR),
        int(ExceptionType.COMPANY_DUPLICATE_ERROR),
    }:
        return True
    return exception_id_int == int(ExceptionType.FORMAT_ERROR) and str(field_name or "").strip() == "company_detail"


def _safe_sheet_title(title: str) -> str:
    cleaned = "".join("_" if char in '\\/:*?[]' else char for char in str(title or "").strip())
    return cleaned[:31] or "Sheet"


def _normalize_file_stem(file_name: str) -> str:
    stem = Path(str(file_name or "异常记录")).stem.strip() or "异常记录"
    return "".join("_" if char in '\\/:*?[]' else char for char in stem)


def _format_company_amount_text(direction: object, profit_loss: object, profit_loss_unit: object) -> str:
    amount = "" if profit_loss is None else str(profit_loss).strip()
    unit = str(profit_loss_unit or "").strip()
    direction_text = ""
    try:
        direction_value = int(direction) if direction is not None else 0
    except (TypeError, ValueError):
        direction_value = 0
    if direction_value > 0:
        direction_text = "增加"
    elif direction_value < 0:
        direction_text = "减少"
    if not amount:
        return ""
    return f"{direction_text}{amount}{unit}".strip() if direction_text else f"{amount}{unit}".strip()


def _format_excel_row_indexes(rows: list[dict[str, Any]]) -> str:
    row_indexes = sorted(
        {
            int(row.get("excel_row_index") or 0)
            for row in rows
            if row.get("excel_row_index") is not None and int(row.get("excel_row_index") or 0) > 0
        }
    )
    return " | ".join(str(value) for value in row_indexes)


def _join_unique(rows: list[dict[str, Any]], field_name: str) -> str:
    values: list[str] = []
    seen: set[str] = set()
    for row in rows:
        value = str(row.get(field_name) or "").strip()
        if not value or value in seen:
            continue
        seen.add(value)
        values.append(value)
    return " | ".join(values)


def _is_punctuation_token(value: object) -> bool:
    text = str(value or "")
    if len(text) != 1:
        return False
    return unicodedata.category(text).startswith("P")


def _build_format_exception_reason(field_name: object, value: object) -> str:
    field = str(field_name or "").strip()
    text = str(value or "").strip()
    if field == "profit_loss_unit":
        return f"金额单位({text})" if text else "金额单位"
    if field in {"main_sentence", "company_detail"} and _is_punctuation_token(text):
        return "标点符号"
    if field == "main_sentence":
        return "主句格式"
    if field == "company_detail":
        return f"{text}格式" if text else "格式"
    if field in {"company_marker", "company_detail_tail"}:
        return "明细段格式"
    return "主句格式"


def _merge_format_exception_details(exception_details: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets: dict[int, dict[str, Any]] = {}
    for item in exception_details:
        if _display_exception_name(item.get("exception_name")) != "格式异常":
            continue
        word_record_id = int(item.get("word_record_id") or 0)
        if not word_record_id:
            continue
        bucket = buckets.setdefault(
            word_record_id,
            {
                "first": item,
                "reasons": [],
            },
        )
        reason = _build_format_exception_reason(item.get("field_name"), item.get("value"))
        if reason and reason not in bucket["reasons"]:
            bucket["reasons"].append(reason)

    merged: list[dict[str, Any]] = []
    seen: set[int] = set()
    for item in exception_details:
        if _display_exception_name(item.get("exception_name")) != "格式异常":
            merged.append(item)
            continue
        word_record_id = int(item.get("word_record_id") or 0)
        if not word_record_id or word_record_id in seen:
            continue
        seen.add(word_record_id)
        bucket = buckets.get(word_record_id)
        if not bucket:
            continue
        first = dict(bucket["first"] or {})
        first["exception_id"] = int(ExceptionType.FORMAT_ERROR)
        first["exception_name"] = "格式异常"
        first["field_name"] = "format_reason"
        first["value"] = ",".join(bucket.get("reasons") or [])
        merged.append(first)
    return merged


def _expand_company_exception_details(
    exception_details: list[dict[str, Any]],
    company_details_cache: dict[int, list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    expanded: list[dict[str, Any]] = []
    for item in exception_details:
        if not _is_company_exception(item.get("exception_id"), item.get("field_name")):
            expanded.append(item)
            continue
        word_record_id = int(item.get("word_record_id") or 0)
        if not word_record_id:
            expanded.append(item)
            continue
        company_name = str(item.get("value") or "").strip()
        if not company_name:
            expanded.append(item)
            continue
        company_details = [
            detail
            for detail in company_details_cache.get(word_record_id, [])
            if str(detail.get("company") or "").strip() == company_name
        ]
        if not company_details:
            cloned = dict(item)
            cloned["company_amount_text"] = ""
            expanded.append(cloned)
            continue
        for detail in company_details:
            cloned = dict(item)
            cloned["company_amount_text"] = _format_company_amount_text(
                detail.get("direction"),
                detail.get("profit_loss"),
                detail.get("profit_loss_unit", ""),
            )
            expanded.append(cloned)
    return expanded


def _merge_company_name_exception_details(
    exception_details: list[dict[str, Any]],
    rows_by_word_record_id: dict[int, list[dict[str, Any]]],
    *,
    exception_ids: set[int],
) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    seen: set[tuple[int, int, str, str, str]] = set()
    for item in exception_details:
        exception_id = int(item.get("exception_id") or 0)
        if exception_id not in exception_ids:
            merged.append(item)
            continue
        word_record_id = int(item.get("word_record_id") or 0)
        first_row = rows_by_word_record_id.get(word_record_id, [{}])[0]
        sheet = str(first_row.get("word_sheet") or "")
        code = str(first_row.get("word_code") or "")
        company = str(item.get("value") or "").strip()
        key = (exception_id, word_record_id, sheet, code, company)
        if key in seen:
            continue
        seen.add(key)
        merged.append(item)
    return merged



def _set_cell_style(cell, *, bold: bool = False, wrap: bool = False) -> None:
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)


def _append_rows(sheet, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        _set_cell_style(cell, bold=True, wrap=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{chr(ord('A') + len(headers) - 1)}{max(1, len(rows) + 1)}"
    for row in rows:
        sheet.append(row)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _fit_columns(sheet, widths: dict[int, int] | None = None) -> None:
    custom_widths = widths or {}
    for index, column_cells in enumerate(sheet.columns, start=1):
        column_letter = column_cells[0].column_letter
        if index in custom_widths:
            sheet.column_dimensions[column_letter].width = custom_widths[index]
            continue
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)


def _build_summary_sheet(
    workbook: Workbook,
    *,
    word_file_name: str,
    excel_file_name: str,
    exception_details: list[dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet(_safe_sheet_title("异常汇总"))
    summary_rows = [
        ("下载时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Word 文件", word_file_name),
        ("Excel 文件", excel_file_name or "-"),
        ("异常总数", len(exception_details)),
    ]
    for label, value in summary_rows:
        sheet.append([label, value])
    for row_index in range(1, len(summary_rows) + 1):
        _set_cell_style(sheet.cell(row=row_index, column=1), bold=True)
        sheet.cell(row=row_index, column=2).alignment = Alignment(vertical="top", wrap_text=True)

    start_row = len(summary_rows) + 2
    headers = ["异常类型", "数量"]
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=column_index, value=header)
        _set_cell_style(cell, bold=True)
    counter = Counter(_display_exception_name(item.get("exception_name")) for item in exception_details)
    for offset, (name, count) in enumerate(sorted(counter.items(), key=lambda item: item[0]), start=1):
        sheet.cell(row=start_row + offset, column=1, value=name).alignment = Alignment(vertical="top")
        sheet.cell(row=start_row + offset, column=2, value=count).alignment = Alignment(vertical="top")
    sheet.freeze_panes = f"A{start_row + 1}"
    end_row = start_row + max(len(counter), 1)
    sheet.auto_filter.ref = f"A{start_row}:B{end_row}"
    _fit_columns(sheet, widths={1: 18, 2: 28})


def _build_detail_rows(
    exception_details: list[dict[str, Any]],
    rows_by_word_record_id: dict[int, list[dict[str, Any]]],
    company_details_cache: dict[int, list[dict[str, Any]]],
    word_file_name: str,
    excel_file_name: str,
) -> list[list[Any]]:
    detail_rows: list[list[Any]] = []
    for index, item in enumerate(exception_details, start=1):
        word_record_id = int(item.get("word_record_id") or 0)
        related_rows = rows_by_word_record_id.get(word_record_id, [])
        matched_excel_rows = [row for row in related_rows if row.get("excel_record_id") is not None]
        first_row = related_rows[0] if related_rows else {}
        company_name = (
            str(item.get("value") or "").strip()
            if _is_company_exception(item.get("exception_id"), item.get("field_name"))
            else ""
        )
        company_detail = None
        if company_name:
            company_detail = next(
                (
                    detail
                    for detail in company_details_cache.get(word_record_id, [])
                    if str(detail.get("company") or "").strip() == company_name
                ),
                None,
            )
        detail_rows.append(
            [
                index,
                _display_exception_name(item.get("exception_name")),
                word_file_name,
                excel_file_name or first_row.get("excel_file_name") or "",
                first_row.get("word_sheet") or "",
                first_row.get("word_code") or "",
                first_row.get("word_name") or "",
                item.get("value") or "",
                company_name,
                item.get("company_amount_text")
                or _format_company_amount_text(
                    company_detail.get("direction") if company_detail else None,
                    company_detail.get("profit_loss") if company_detail else None,
                    company_detail.get("profit_loss_unit") if company_detail else "",
                ),
                len(matched_excel_rows),
                _join_unique(matched_excel_rows, "excel_sheet"),
                _join_unique(matched_excel_rows, "excel_code"),
                _join_unique(matched_excel_rows, "excel_name"),
                _format_excel_row_indexes(matched_excel_rows),
                first_row.get("word_context") or "",
            ]
        )
    return detail_rows


def _build_company_exception_rows(
    exception_details: list[dict[str, Any]],
    rows_by_word_record_id: dict[int, list[dict[str, Any]]],
    company_details_cache: dict[int, list[dict[str, Any]]],
    word_file_name: str,
    excel_file_name: str,
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for item in exception_details:
        if not _is_company_exception(item.get("exception_id"), item.get("field_name")):
            continue
        word_record_id = int(item.get("word_record_id") or 0)
        company_name = str(item.get("value") or "").strip()
        related_rows = rows_by_word_record_id.get(word_record_id, [])
        matched_excel_rows = [row for row in related_rows if row.get("excel_record_id") is not None]
        first_row = related_rows[0] if related_rows else {}
        company_detail = next(
            (
                detail
                for detail in company_details_cache.get(word_record_id, [])
                if str(detail.get("company") or "").strip() == company_name
            ),
            None,
        )
        rows.append(
            [
                len(rows) + 1,
                company_name,
                item.get("company_amount_text")
                or _format_company_amount_text(
                    company_detail.get("direction") if company_detail else None,
                    company_detail.get("profit_loss") if company_detail else None,
                    company_detail.get("profit_loss_unit") if company_detail else "",
                ),
                word_file_name,
                excel_file_name or first_row.get("excel_file_name") or "",
                first_row.get("word_sheet") or "",
                first_row.get("word_code") or "",
                first_row.get("word_name") or "",
                len(matched_excel_rows),
                _format_excel_row_indexes(matched_excel_rows),
                first_row.get("word_context") or "",
            ]
        )
    return rows


def _build_multi_excel_rows(
    exception_details: list[dict[str, Any]],
    rows_by_word_record_id: dict[int, list[dict[str, Any]]],
    word_file_name: str,
    excel_file_name: str,
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    seen_word_record_ids: set[int] = set()
    for item in exception_details:
        if int(item.get("exception_id") or 0) != int(ExceptionType.EXCEL_ERROR):
            continue
        word_record_id = int(item.get("word_record_id") or 0)
        if word_record_id in seen_word_record_ids:
            continue
        seen_word_record_ids.add(word_record_id)
        related_rows = [row for row in rows_by_word_record_id.get(word_record_id, []) if row.get("excel_record_id") is not None]
        first_row = related_rows[0] if related_rows else (rows_by_word_record_id.get(word_record_id, [{}])[0])
        for candidate_index, candidate in enumerate(related_rows, start=1):
            rows.append(
                [
                    len(rows) + 1,
                    word_file_name,
                    excel_file_name or candidate.get("excel_file_name") or "",
                    word_record_id,
                    first_row.get("word_sheet") or "",
                    first_row.get("word_code") or "",
                    first_row.get("word_name") or "",
                    candidate_index,
                    candidate.get("excel_sheet") or "",
                    candidate.get("excel_code") or "",
                    candidate.get("excel_name") or "",
                    candidate.get("excel_row_index") or "",
                    first_row.get("word_context") or "",
                ]
            )
    return rows


async def export_document_pair_exceptions(
    db: AsyncSession,
    batch_id: str,
    word_file_name: str,
    excel_file_name: str = "",
) -> tuple[str, bytes]:
    """导出指定 Word 文件对的异常记录为 xlsx，返回 (文件名, 字节内容)。"""

    # 从任务记录解析文件名（兜底）。
    task = (await db.execute(select(CreditCompareTask).where(CreditCompareTask.batch_id == batch_id))).scalars().first()
    resolved_word_file = str(word_file_name or (task.word_file_name if task else "") or "").strip()
    resolved_excel_file = str(excel_file_name or (task.excel_file_name if task else "") or "").strip()
    if not batch_id or not resolved_word_file:
        raise FileNotFoundError("缺少批次号或 Word 文件名，无法导出异常记录。")

    pair_rows = [
        row
        for row in await repository.list_compare_link_details(db, batch_id)
        if str(row.get("word_file_name") or "") == resolved_word_file
    ]
    if not pair_rows:
        raise FileNotFoundError("未找到对应详情记录，无法导出异常记录。")

    rows_by_word_record_id: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in pair_rows:
        rows_by_word_record_id[int(row.get("word_record_id") or 0)].append(row)
    company_details_cache: dict[int, list[dict[str, Any]]] = {}
    for word_record_id in rows_by_word_record_id:
        company_details_cache[word_record_id] = await repository.list_company_by_word_record(db, word_record_id)

    target_word_record_ids = set(rows_by_word_record_id.keys())
    exception_details = [
        item
        for item in await repository.list_exception_details_by_batch(db, batch_id)
        if int(item.get("word_record_id") or 0) in target_word_record_ids
    ]
    export_exception_details = _merge_format_exception_details(exception_details)
    export_exception_details = _expand_company_exception_details(export_exception_details, company_details_cache)
    export_exception_details = _merge_company_name_exception_details(
        export_exception_details,
        rows_by_word_record_id,
        exception_ids={int(ExceptionType.COMPANY_DIRECTION_ERROR), int(ExceptionType.COMPANY_DUPLICATE_ERROR)},
    )

    workbook = Workbook()
    workbook.remove(workbook.active)

    _build_summary_sheet(
        workbook,
        word_file_name=resolved_word_file,
        excel_file_name=resolved_excel_file,
        exception_details=export_exception_details,
    )

    detail_sheet = workbook.create_sheet(_safe_sheet_title("异常明细"))
    detail_headers = [
        "序号", "异常类型", "Word文件", "Excel文件", "表单", "指标代码", "指标名称", "异常值",
        "关联公司", "公司增减金额", "Excel候选数", "Excel表单", "Excel指标代码", "Excel指标名称",
        "Excel行号", "Word原文",
    ]
    _append_rows(
        detail_sheet,
        detail_headers,
        _build_detail_rows(
            export_exception_details, rows_by_word_record_id, company_details_cache, resolved_word_file, resolved_excel_file
        ),
    )
    _fit_columns(detail_sheet, widths={2: 22, 3: 24, 4: 24, 5: 12, 6: 14, 7: 20, 8: 22, 9: 20, 10: 18, 11: 12, 12: 12, 13: 16, 14: 20, 15: 14, 16: 50})

    company_sheet = workbook.create_sheet(_safe_sheet_title("关联公司增减金额异常"))
    company_headers = [
        "序号", "公司名称", "公司增减金额", "Word文件", "Excel文件", "表单", "指标代码", "指标名称",
        "Excel候选数", "Excel行号", "Word原文",
    ]
    _append_rows(
        company_sheet,
        company_headers,
        _build_company_exception_rows(
            export_exception_details, rows_by_word_record_id, company_details_cache, resolved_word_file, resolved_excel_file
        ),
    )
    _fit_columns(company_sheet, widths={2: 28, 3: 18, 4: 24, 5: 24, 6: 12, 7: 14, 8: 20, 9: 12, 10: 14, 11: 50})

    multi_excel_sheet = workbook.create_sheet(_safe_sheet_title("多Excel候选"))
    multi_excel_headers = [
        "序号", "Word文件", "Excel文件", "Word记录ID", "表单", "指标代码", "指标名称", "候选序号",
        "Excel表单", "Excel指标代码", "Excel指标名称", "Excel行号", "Word原文",
    ]
    _append_rows(
        multi_excel_sheet,
        multi_excel_headers,
        _build_multi_excel_rows(export_exception_details, rows_by_word_record_id, resolved_word_file, resolved_excel_file),
    )
    _fit_columns(multi_excel_sheet, widths={2: 24, 3: 24, 5: 12, 6: 14, 7: 20, 8: 10, 9: 12, 10: 16, 11: 20, 12: 12, 13: 50})

    buffer = BytesIO()
    workbook.save(buffer)
    file_name = f"{_normalize_file_stem(resolved_word_file)}_异常记录.xlsx"
    return file_name, buffer.getvalue()
