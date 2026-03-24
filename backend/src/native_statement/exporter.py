"""
Excel 导出器

将解析结果导出为格式化的 Excel 文件。
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# 样式定义
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=False)
SUMMARY_LABEL_FONT = Font(bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def export_to_excel(result: dict, output_path: str = None) -> str | io.BytesIO:
    """
    将解析结果导出为 Excel 文件

    Args:
        result: parse_native_pdf 返回的结果字典
        output_path: 输出文件路径。如果为 None，返回 BytesIO 对象

    Returns:
        文件路径或 BytesIO 对象
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "银行流水"

    current_row = 1

    # ============================================================
    # 写入汇总信息
    # ============================================================
    summary = result.get("summary", {})
    if summary:
        # 标题
        ws.cell(row=current_row, column=1, value="汇总信息").font = Font(bold=True, size=13)
        current_row += 1

        summary_labels = {
            "account_name": "户名",
            "account_number": "账号",
            "bank_name": "开户行",
            "currency": "币种",
            "date_range": "起止日期",
            "start_date": "起始日期",
            "end_date": "结束日期",
            "income_total": "收入总金额",
            "expense_total": "支出总金额",
            "income_count": "收入总笔数",
            "expense_count": "支出总笔数",
        }

        for field, label in summary_labels.items():
            value = summary.get(field, "")
            if value:
                ws.cell(row=current_row, column=1, value=label).font = SUMMARY_LABEL_FONT
                
                # 金额汇总字段转为数值格式
                if field in ["income_total", "expense_total"]:
                    try:
                        num_val = float(str(value).replace(",", "").strip())
                        cell = ws.cell(row=current_row, column=2, value=num_val)
                        cell.number_format = '0.00'
                    except ValueError:
                        ws.cell(row=current_row, column=2, value=value)
                else:
                    ws.cell(row=current_row, column=2, value=value)
                
                current_row += 1

        # 空行分隔
        current_row += 1

    # ============================================================
    # 写入交易明细
    # ============================================================
    transactions = result.get("transactions", [])
    raw_headers = result.get("raw_headers", [])
    mapped_headers = result.get("headers", [])

    if transactions:
        # 标题
        ws.cell(row=current_row, column=1, value="交易明细").font = Font(bold=True, size=13)
        ws.cell(row=current_row, column=2, value=f"共 {len(transactions)} 条记录")
        current_row += 1

        # 表头（使用原始中文表头）
        display_headers = raw_headers if raw_headers else mapped_headers
        for col_idx, header in enumerate(display_headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
        current_row += 1

        # 数据行
        for tx in transactions:
            for col_idx, field in enumerate(mapped_headers, 1):
                value = tx.get(field, "")
                
                # 针对金额字段尝试转换为数字并设置格式
                is_num = False
                if field in ["income", "expense", "balance", "收入", "支出", "账户余额", "借方发生额", "贷方发生额", "余额", "存入金额", "支出金额", "收入金额"]:
                    try:
                        # 有的数值可能有千分位占位符，需剔除
                        num_val = float(str(value).replace(",", "").strip())
                        value = num_val
                        is_num = True
                    except ValueError:
                        pass

                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER
                if is_num:
                    cell.number_format = '0.00'
                    
            current_row += 1

        # 自动调整列宽
        for col_idx in range(1, len(display_headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_width = len(str(display_headers[col_idx - 1])) * 2  # 中文字符宽度
            for row in ws.iter_rows(min_row=current_row - len(transactions),
                                     max_row=current_row - 1,
                                     min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_width = max(max_width, len(str(cell.value)) * 1.2)
            ws.column_dimensions[col_letter].width = min(max_width + 2, 40)

    # 写入元信息
    ws2 = wb.create_sheet("解析信息")
    ws2.cell(row=1, column=1, value="银行类型").font = SUMMARY_LABEL_FONT
    ws2.cell(row=1, column=2, value=result.get("bank_type", "unknown"))
    ws2.cell(row=2, column=1, value="页数").font = SUMMARY_LABEL_FONT
    ws2.cell(row=2, column=2, value=result.get("page_count", 0))
    ws2.cell(row=3, column=1, value="交易记录数").font = SUMMARY_LABEL_FONT
    ws2.cell(row=3, column=2, value=result.get("total_rows", 0))

    # 输出
    if output_path:
        wb.save(output_path)
        return output_path
    else:
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
